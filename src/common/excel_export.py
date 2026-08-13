from __future__ import annotations

import io
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import case_store
from check_result import CheckResult, DISPLAY_LABEL

HEADER_FILL = PatternFill(start_color="12161F", end_color="12161F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="0F1420")
SECTION_FONT = Font(bold=True, size=11, color="0F1420")
LABEL_FONT = Font(bold=True, size=9, color="6B7280")
STATUS_FILL = {
    "PASS": PatternFill(start_color="E7F6ED", end_color="E7F6ED", fill_type="solid"),
    "FAIL": PatternFill(start_color="FCEAE9", end_color="FCEAE9", fill_type="solid"),
    "REVIEW": PatternFill(start_color="FBF1DE", end_color="FBF1DE", fill_type="solid"),
    "N/A": PatternFill(start_color="EEF0F3", end_color="EEF0F3", fill_type="solid"),
}
STATUS_FONT = {
    "PASS": Font(color="146C3A", bold=True),
    "FAIL": Font(color="B3261E", bold=True),
    "REVIEW": Font(color="91600A", bold=True),
    "N/A": Font(color="5B6472", bold=True),
}
WRAP = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _exceptions_block(ws, row: int, title: str, results: list[CheckResult], side: str) -> int:
    value_attr = f"{side}_value"
    exceptions = [r for r in results if r.status in ("FAIL", "REVIEW")]

    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=10.5, color="0F1420")
    row += 1

    if not exceptions:
        ws.cell(row=row, column=1, value="No exceptions.").font = Font(italic=True, size=9, color="6B7280")
        return row + 2

    headers = ["KCT", "Check", "Status", "Confidence", "Value", "Note"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for r in exceptions:
        values = [
            r.kct, r.check, DISPLAY_LABEL.get(r.status, r.status),
            f"{r.confidence:.0f}%" if r.confidence is not None else "",
            getattr(r, value_attr), r.note,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)
        status_cell = ws.cell(row=row, column=3)
        status_cell.fill = STATUS_FILL.get(r.status, STATUS_FILL["N/A"])
        status_cell.font = STATUS_FONT.get(r.status, STATUS_FONT["N/A"])
        for col in (2, 5, 6):
            ws.cell(row=row, column=col).alignment = WRAP
        row += 1

    return row + 1


_REMARKS_SECTION_TITLES = [
    ("executive_summary", "EXECUTIVE SUMMARY"),
    ("positive_indicators", "POSITIVE INDICATORS"),
    ("areas_of_concern", "AREAS OF CONCERN"),
    ("recommendations", "AI RECOMMENDATIONS"),
]


def _format_remarks(remarks: str) -> str:
    if not remarks:
        return "Not generated."
    try:
        sections = json.loads(remarks)
        if not isinstance(sections, dict):
            raise ValueError
    except (json.JSONDecodeError, ValueError, TypeError):
        # Older cases stored remarks as a plain bullet-list string -- keep as-is.
        return remarks
    blocks = []
    for key, title in _REMARKS_SECTION_TITLES:
        bullets = sections.get(key) or []
        if bullets:
            blocks.append(title + ":\n" + "\n".join(f"- {b}" for b in bullets))
    return "\n\n".join(blocks) or "No AI summary available."


_AI_SUMMARY_SECTIONS = [
    ("executive_summary", "Executive Summary", "FDEBD3"),
    ("positive_indicators", "Positive Indicators", "D9EAD3"),
    ("areas_of_concern", "Areas of Concern", "FCE0B4"),
    ("recommendations", "AI Recommendations", "D6E4F0"),
]


def _ai_summary_block(ws, row: int, remarks: str) -> int:
    """Renders the AI summary as one colored section per category (matching the reference
    layout), each with its own header row and one row per bullet -- not a single merged
    text block."""
    ws.cell(row=row, column=1, value="AI Summary").font = SECTION_FONT
    row += 2

    try:
        sections = json.loads(remarks) if remarks else None
        if sections is not None and not isinstance(sections, dict):
            raise ValueError
    except (json.JSONDecodeError, ValueError, TypeError):
        sections = None

    if sections is None:
        # Older cases stored remarks as a plain bullet-list string -- render as-is.
        text = remarks or "Not generated."
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = max(60, 16 * (text.count("\n") + 2))
        return row + 2

    any_section = False
    for key, title, color in _AI_SUMMARY_SECTIONS:
        bullets = [b for b in (sections.get(key) or []) if b]
        if not bullets:
            continue
        any_section = True
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=10.5, color="0F1420")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill
        row += 1
        for bullet in bullets:
            cell = ws.cell(row=row, column=2, value=bullet)
            cell.alignment = WRAP
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill
            row += 1

    if not any_section:
        ws.cell(row=row, column=1, value="No AI summary available.").font = Font(italic=True, size=9, color="6B7280")
        row += 1
    return row + 1


def _full_report_sheet(
    wb: Workbook,
    case_meta: dict,
    results: list[CheckResult],
    remarks: str,
    left_label: str,
    right_label: str,
) -> None:
    ws = wb.create_sheet("Full Report", 0)
    row = 1

    ws.cell(row=row, column=1, value=case_meta.get("title", "Control Testing Report")).font = TITLE_FONT
    row += 2

    meta_fields = [
        ("Case ID", case_meta.get("case_id", "")),
        ("Module", case_meta.get("module", "")),
        ("Processed At", case_meta.get("processed_at", "")),
        ("Processing Time", case_meta.get("processing_time", "")),
        ("Documents", case_meta.get("documents", "")),
        ("Overall Status", case_meta.get("overall_status", "")),
    ]
    for label, value in meta_fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=str(value))
        cell.alignment = WRAP
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Exceptions").font = SECTION_FONT
    row += 2

    row = _exceptions_block(ws, row, left_label, results, "left")
    row = _exceptions_block(ws, row, right_label, results, "right")

    row = _ai_summary_block(ws, row, remarks)

    _autosize(ws, {1: 16, 2: 30, 3: 10, 4: 12, 5: 34, 6: 48})


def build_workbook(
    case_meta: dict,
    results: list[CheckResult],
    remarks: str,
    left_label: str,
    right_label: str,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _full_report_sheet(wb, case_meta, results, remarks, left_label, right_label)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


CASE_STATUS_LABEL = {"complete": "Complete", "needs_review": "Needs Review", "error": "Error"}
_CASE_STATUS_TO_CHECK_STATUS = {"complete": "PASS", "needs_review": "REVIEW", "error": "FAIL"}


def _format_duration(seconds) -> str:
    total = int(seconds or 0)
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def build_consolidated_workbook(
    case_type: str, module_name: str, case_ids: list[str] | None = None, ai_summary: str = ""
) -> bytes:
    df = case_store.list_cases(case_type)
    if case_ids is not None:
        df = df[df["case_id"].isin(case_ids)]
    wb = Workbook()
    wb.properties.title = f"{module_name} Cases"
    ws = wb.active
    ws.title = "Cases"
    headers = [
        "Case ID", "Customer", "Completed At", "Processing Time", "Status", "Accuracy",
        "Match", "Mismatch", "Review", "N/A", "Remarks",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for _, row in df.iterrows():
        case_status = row.get("status") or ("needs_review" if row["flagged"] else "complete")
        accuracy = row.get("accuracy")
        ws.append([
            row["case_id"], row.get("customer_name") or "", row["created_at"],
            _format_duration(row["processing_seconds"]), CASE_STATUS_LABEL.get(case_status, case_status),
            f"{accuracy:.0f}%" if pd.notna(accuracy) else "—",
            row["pass_count"], row["fail_count"], row["review_count"], row["na_count"], _format_remarks(row["remarks"]),
        ])
        status_cell = ws.cell(row=ws.max_row, column=5)
        check_status = _CASE_STATUS_TO_CHECK_STATUS.get(case_status, "N/A")
        status_cell.fill = STATUS_FILL[check_status]
        status_cell.font = STATUS_FONT[check_status]
        ws.cell(row=ws.max_row, column=11).alignment = WRAP
    _autosize(ws, {1: 26, 2: 22, 3: 20, 4: 14, 5: 14, 6: 10, 7: 8, 8: 8, 9: 8, 10: 8, 11: 60})

    ws2 = wb.create_sheet("Exceptions")
    headers2 = ["Case ID", "KCT", "Check", "Status", "Confidence", "Note"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws2.freeze_panes = "A2"
    for _, row in df.iterrows():
        for r in json.loads(row["results_json"]):
            if r.get("status") not in ("FAIL", "REVIEW"):
                continue
            ws2.append([
                row["case_id"], r.get("kct"), r.get("check"), DISPLAY_LABEL.get(r.get("status"), r.get("status")),
                f"{r['confidence']:.0f}%" if r.get("confidence") is not None else "",
                r.get("note"),
            ])
            status_cell = ws2.cell(row=ws2.max_row, column=4)
            status_cell.fill = STATUS_FILL.get(r.get("status"), STATUS_FILL["N/A"])
            status_cell.font = STATUS_FONT.get(r.get("status"), STATUS_FONT["N/A"])
            ws2.cell(row=ws2.max_row, column=6).alignment = WRAP
    _autosize(ws2, {1: 26, 2: 14, 3: 34, 4: 10, 5: 12, 6: 48})

    ws3 = wb.create_sheet("AI Summary")
    ws3.cell(row=1, column=1, value=f"{module_name} — AI Executive Summary").font = TITLE_FONT
    text = (ai_summary or "").strip() or (
        'Not yet generated -- click "Generate / regenerate final report" on the Dashboard '
        "or Reports page, then re-export."
    )
    ws3.cell(row=3, column=1, value=text)
    ws3.cell(row=3, column=1).alignment = WRAP
    ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    ws3.row_dimensions[3].height = max(80, 16 * (text.count("\n") + 2))
    _autosize(ws3, {1: 100})

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
